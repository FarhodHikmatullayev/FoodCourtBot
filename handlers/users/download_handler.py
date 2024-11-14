import os
import tempfile

import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from openpyxl.styles import Alignment

from keyboards.default.menu_keyboards import back_to_menu
from keyboards.inline.categories import category_keyboard, category_callback_data
from loader import db, dp
from states.table_states import TableState


async def download_food_comments(days):
    food_comments = await db.select_recent_food_comments(days=days)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'T/r'
    worksheet['B1'] = "Stol raqami"
    worksheet['C1'] = "Baho"
    worksheet['D1'] = 'Vaqt'

    for cell in ['A1', 'B1', 'C1', 'D1']:
        worksheet[cell].alignment = Alignment(horizontal='center')

    worksheet.cell(row=1, column=1, value='T/r')
    worksheet.cell(row=1, column=2, value='Stol raqami')
    worksheet.cell(row=1, column=3, value="Baho")
    worksheet.cell(row=1, column=4, value='Vaqt')

    tr = 0
    for food_comment in food_comments:
        tr += 1

        table_number = food_comment['table_number']
        grade = food_comment['grade']
        created_at = food_comment['created_at']

        worksheet.cell(row=tr + 1, column=1, value=tr).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=2, value=table_number).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=3, value=grade).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=4, value=created_at.strftime("%d.%m.%Y  %H:%M")).alignment = Alignment(
            horizontal='center')

    temp_dir = tempfile.gettempdir()
    file_name = f"FoodCommentsFor_{days}days.xlsx"
    file_path = os.path.join(temp_dir, file_name)
    workbook.save(file_path)

    return temp_dir


async def download_waiter_comments(days):
    waiter_comments = await db.select_recent_waiter_comments(days=days)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'T/r'
    worksheet['B1'] = "Stol raqami"
    worksheet['C1'] = "Baho"
    worksheet['D1'] = 'Vaqt'

    for cell in ['A1', 'B1', 'C1', 'D1']:
        worksheet[cell].alignment = Alignment(horizontal='center')

    worksheet.cell(row=1, column=1, value='T/r')
    worksheet.cell(row=1, column=2, value='Stol raqami')
    worksheet.cell(row=1, column=3, value="Baho")
    worksheet.cell(row=1, column=4, value='Vaqt')

    tr = 0
    for waiter_comment in waiter_comments:
        tr += 1

        table_number = waiter_comment['table_number']
        grade = waiter_comment['grade']
        created_at = waiter_comment['created_at']

        worksheet.cell(row=tr + 1, column=1, value=tr).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=2, value=table_number).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=3, value=grade).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=4, value=created_at.strftime("%d.%m.%Y  %H:%M")).alignment = Alignment(
            horizontal='center')

    temp_dir = tempfile.gettempdir()
    file_name = f"WaiterCommentsFor_{days}days.xlsx"
    file_path = os.path.join(temp_dir, file_name)
    workbook.save(file_path)

    return temp_dir


async def download_comments(days):
    comments = await db.select_recent_comments(days=days)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'T/r'
    worksheet['B1'] = "Vaqt"
    worksheet['C1'] = "Stol raqami"
    worksheet['D1'] = "Fikr/Taklif"

    for cell in ['A1', 'B1', 'C1', 'D1']:
        worksheet[cell].alignment = Alignment(horizontal='center')

    worksheet.cell(row=1, column=1, value='T/r')
    worksheet.cell(row=1, column=2, value='Vaqt')
    worksheet.cell(row=1, column=3, value='Stol raqami')
    worksheet.cell(row=1, column=4, value="Fikr/Taklif")

    tr = 0
    for comment in comments:
        tr += 1
        print('comment', comment)

        table_number = comment['table_number']
        comment1 = comment['comment']
        created_at = comment['created_at']

        worksheet.cell(row=tr + 1, column=1, value=tr).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=2, value=created_at.strftime("%d.%m.%Y  %H:%M")).alignment = Alignment(
            horizontal='center')
        worksheet.cell(row=tr + 1, column=3, value=table_number).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=4, value=comment1).alignment = Alignment(
            horizontal='center')

    temp_dir = tempfile.gettempdir()
    file_name = f"CommentsFor_{days}days.xlsx"
    file_path = os.path.join(temp_dir, file_name)
    workbook.save(file_path)

    return temp_dir


@dp.callback_query_handler(text="download", state="*")
async def get_category(call: types.CallbackQuery, state: FSMContext):
    await TableState.table_number.set()
    await call.message.edit_text(text="📂 Yuklab olmoqchi bo'lgan kategoriyangizni tanlang:\n",
                              reply_markup=category_keyboard)


@dp.callback_query_handler(category_callback_data.filter(), state="*")
async def get_days(call: types.CallbackQuery, state: FSMContext, callback_data: dict):
    print('callback_data', callback_data)
    category = callback_data.get('category')
    await state.update_data(category=category)
    await call.message.edit_text(text="📅 Necha kunlik natijalarni yuklab olmoqchisiz?\n"
                                   "Misol: 20")
    await TableState.days.set()


@dp.message_handler(state=TableState.days)
async def download_function(message: types.Message, state: FSMContext):
    days = message.text
    try:
        days = int(days)
    except:
        await message.answer(text="📅 Siz kun uchun son kiritishingiz kerak:", reply_markup=back_to_menu)
        return
    await state.update_data(days=days)

    data = await state.get_data()
    print('data', data)
    category = data.get("category")

    if category == 'waiter':
        temp_dir = await download_waiter_comments(days=days)
        file_name = f"WaiterCommentsFor_{days}days.xlsx"
        with open(os.path.join(temp_dir, file_name), 'rb') as file:
            await message.answer_document(document=file)
        os.remove(os.path.join(temp_dir, file_name))

    elif category == 'food':
        temp_dir = await download_food_comments(days=days)
        file_name = f"FoodCommentsFor_{days}days.xlsx"
        with open(os.path.join(temp_dir, file_name), 'rb') as file:
            await message.answer_document(document=file)
        os.remove(os.path.join(temp_dir, file_name))

    else:
        temp_dir = await download_comments(days=days)
        file_name = f"CommentsFor_{days}days.xlsx"
        with open(os.path.join(temp_dir, file_name), 'rb') as file:
            await message.answer_document(document=file)
        os.remove(os.path.join(temp_dir, file_name))
